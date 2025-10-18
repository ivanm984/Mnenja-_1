"""Generation of filled-in Excel forms (Priloga 10A)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

try:  # pragma: no cover - optional dependency import guard
    from openpyxl import load_workbook
    from openpyxl.utils.cell import coordinate_to_tuple
    from openpyxl.styles import Alignment
except ImportError as exc:  # pragma: no cover - import guard
    raise RuntimeError(
        "Knjižnica 'openpyxl' ni nameščena. Namestite jo z `pip install openpyxl`."
    ) from exc

from .config import PROJECT_ROOT

TEMPLATE_PATH = PROJECT_ROOT / "Priloga10A.xlsx"

KEY_DATA_LABELS = [
    ("glavni_objekt", "Glavni objekt"),
    ("vrsta_gradnje", "Vrsta gradnje"),
    ("klasifikacija_cc_si", "Klasifikacija CC-SI"),
    ("parcela_objekta", "Parcela objekta"),
    ("stevilke_parcel_ko", "Parcele in k.o."),
    ("velikost_parcel", "Velikost parcel"),
    ("velikost_obstojecega_objekta", "Obstoječi objekt"),
    ("tlorisne_dimenzije", "Dimenzije novega objekta"),
    ("gabariti_etaznost", "Gabariti in etažnost"),
    ("faktor_zazidanosti_fz", "Faktor zazidanosti"),
    ("faktor_izrabe_fi", "Faktor izrabe"),
    ("zelene_povrsine", "Zelene površine"),
    ("naklon_strehe", "Naklon strehe"),
    ("kritina_barva", "Kritina in barva"),
    ("materiali_gradnje", "Materiali"),
    ("smer_slemena", "Smer slemena"),
    ("visinske_kote", "Višinske kote"),
    ("odmiki_parcel", "Odmiki"),
    ("komunalni_prikljucki", "Komunalni priključki"),
]


def _clean(value: Any, fallback: str = "Ni podatka") -> str:
    if value is None:
        return fallback
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    return text or fallback


def _format_key_data(key_data: Dict[str, Any]) -> str:
    lines: List[str] = []
    for key, label in KEY_DATA_LABELS:
        value = _clean(key_data.get(key, ""), "")
        if value and value.lower() != "ni podatka v dokumentaciji":
            lines.append(f"• {label}: {value}")
    return "\n".join(lines) or "Ni potrjenih ključnih podatkov iz projekta."


def _format_predpis(zahteve: Iterable[Dict[str, Any]]) -> str:
    seen: List[str] = []
    for zahteva in zahteve:
        naslov = _clean(zahteva.get("naslov", ""), "")
        if naslov and naslov not in seen:
            seen.append(naslov)
    if not seen:
        return "Ni evidentiranih pravnih podlag."
    return "\n".join(seen)


def _format_obrazlozitev(
    total: int, non_compliant: List[str], compliant: List[str], key_data_text: str
) -> str:
    lines = [
        f"Analiziranih pogojev: {total}",
        f"Neskladnih pogojev: {len(non_compliant)}",
        f"Skladnih pogojev: {len(compliant)}",
    ]

    if non_compliant:
        lines.append("")
        lines.append("Neskladni členi:")
        lines.extend([f"  • {item}" for item in non_compliant])

    if compliant:
        lines.append("")
        lines.append("Skladni členi:")
        lines.extend([f"  • {item}" for item in compliant])

    cleaned_key_data = key_data_text.strip()
    if cleaned_key_data:
        lines.append("")
        lines.append("Ključni podatki projekta:")
        lines.append(cleaned_key_data)

    return "\n".join(lines)


def _summarize_results(zahteve: Iterable[Dict[str, Any]], results_map: Dict[str, Dict[str, Any]]):
    compliant: List[str] = []
    non_compliant: List[str] = []
    for zahteva in zahteve:
        zid = zahteva.get("id")
        result = results_map.get(zid, {}) if zid else {}
        status = _clean(result.get("skladnost", "Neznano"))
        naslov = _clean(zahteva.get("naslov", "Zahteva"))
        line = f"{naslov} – {status}"
        if "nesklad" in status.lower():
            non_compliant.append(line)
        else:
            compliant.append(line)
    return compliant, non_compliant


def _format_source_files(source_files: Iterable[Dict[str, Any]]) -> str:
    files = []
    for item in source_files:
        name = _clean(item.get("filename", ""), "")
        pages = _clean(item.get("pages", ""), "")
        if pages:
            files.append(f"{name} (strani: {pages})")
        else:
            files.append(name)
    return "\n".join(files) if files else "Ni navedenih dokumentov."


def _apply_wrap_text(cell) -> None:
    try:
        current = cell.alignment
    except AttributeError:
        current = None

    if current:
        cell.alignment = Alignment(
            horizontal=current.horizontal,
            vertical=current.vertical,
            text_rotation=current.text_rotation,
            wrap_text=True,
            shrink_to_fit=current.shrink_to_fit,
            indent=current.indent,
            relative_indent=getattr(current, "relative_indent", 0),
            justify_last_line=getattr(current, "justify_last_line", False),
            reading_order=getattr(current, "reading_order", 0),
        )
    else:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _set_cell_value(worksheet, cell: str, value: Any) -> None:
    try:
        target = worksheet[cell]
        target.value = value
        _apply_wrap_text(target)
        return
    except AttributeError as exc:
        if "MergedCell" not in str(exc):
            raise

        row, column = coordinate_to_tuple(cell)
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= column <= merged_range.max_col
            ):
                base_cell = worksheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                )
                base_cell.value = value
                _apply_wrap_text(base_cell)
                return

        # If the cell is not part of a merged range re-raise the original error.
        raise exc


def generate_priloga_10a(
    zahteve: List[Dict[str, Any]],
    results_map: Dict[str, Dict[str, Any]],
    metadata: Dict[str, Any],
    key_data: Dict[str, Any],
    source_files: Iterable[Dict[str, Any]],
    output_path: str,
) -> str:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Manjka predloga Priloga10A.xlsx na poti: {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook.active

    project_name = _clean(metadata.get("ime_projekta", "Ni podatka"))
    _set_cell_value(
        worksheet,
        "C4",
        f"Mnenje o skladnosti – {project_name}" if project_name else "Mnenje o skladnosti",
    )
    _set_cell_value(worksheet, "C7", _clean(metadata.get("mnenjedajalec", "Avtomatski pregled skladnosti")))
    _set_cell_value(worksheet, "C9", _clean(metadata.get("stevilka_porocila", "Ni podatka")))
    _set_cell_value(worksheet, "C10", datetime.now().strftime("%d.%m.%Y"))
    _set_cell_value(worksheet, "C11", _format_predpis(zahteve))
    _set_cell_value(worksheet, "C12", _clean(metadata.get("postopek_vodil", "Ni podatka")))
    _set_cell_value(worksheet, "C14", _clean(metadata.get("odgovorna_oseba", "Ni podatka")))

    _set_cell_value(worksheet, "C34", project_name)
    _set_cell_value(
        worksheet,
        "C35",
        _clean(metadata.get("kratek_opis", key_data.get("vrsta_gradnje", "Ni podatka"))),
    )
    _set_cell_value(worksheet, "C37", _clean(metadata.get("stevilka_projekta", "Ni podatka")))
    _set_cell_value(worksheet, "C38", _clean(metadata.get("datum_projekta", "Ni podatka")))
    _set_cell_value(worksheet, "C39", _clean(metadata.get("projektant", "Ni podatka")))

    _set_cell_value(worksheet, "C47", _format_source_files(source_files))

    compliant, non_compliant = _summarize_results(zahteve, results_map)
    total = len(zahteve)
    overall_skladnost = "SKLADNA" if not non_compliant else "NESKLADNA"
    _set_cell_value(worksheet, "B48", "X" if overall_skladnost == "SKLADNA" else "")
    _set_cell_value(worksheet, "B49", "X" if overall_skladnost == "NESKLADNA" else "")

    _set_cell_value(worksheet, "C52", "")
    _set_cell_value(worksheet, "C53", "")
    _set_cell_value(worksheet, "C54", "")

    key_data_text = _format_key_data(key_data)
    obrazlozitev_text = _format_obrazlozitev(total, non_compliant, compliant, key_data_text)
    _set_cell_value(worksheet, "C57", obrazlozitev_text)
    _set_cell_value(
        worksheet,
        "C62",
        f"Gradnja je {overall_skladnost.lower()} glede na preverjene pogoje." if total else "Analiza pogojev ni bila izvedena."
    )
    _set_cell_value(worksheet, "C40", _clean(metadata.get("pvo_status", "Ni podatka")))

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    return str(output_file.resolve())


__all__ = ["generate_priloga_10a"]
